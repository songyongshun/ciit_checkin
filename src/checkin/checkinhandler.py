import importlib.resources
from http.server import BaseHTTPRequestHandler
import json
import os
import re
import urllib.parse
import subprocess
import datetime
from .database import (
    get_all_classrooms,
    add_classroom,
    delete_classroom,
    get_classroom_by_id,
    get_class_student_counts,
    delete_students_by_class_name,
    DATABASE_PATH,
)
import qrcode
from PIL import ImageDraw, ImageFont
import sqlite3

class CheckinHandler(BaseHTTPRequestHandler):
    public_ip = "127.0.0.1"  # 将作为实例属性或通过 run_server 设置

    # 全局签到状态字典：classroom_id -> bool (True=允许签到)
    checkin_enabled = {}

    # 内联 admin 页面模板（不再使用外部文件）
    _admin_template = '''<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>姓名记录管理</title>
  <style>
  body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    margin: 0;
    padding: 20px;
    background-color: #f5f5f5;
  }}
  .container {{
    max-width: 800px;
    margin: 0 auto;
    background-color: white;
    padding: 20px;
    border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
  }}
  h2 {{
    text-align: center;
    color: #333;
    margin-bottom: 20px;
  }}
  table {{
    width: 100%;
    border-collapse: collapse;
    margin-top: 10px;
  }}
  td {{
    padding: 8px;
    text-align: center;
    border: 1px solid #ddd;
  }}
  .btn {{
    display: block;
    width: 200px;
    margin: 10px auto;
    padding: 10px;
    background-color: #4CAF50;
    color: white;
    border: none;
    border-radius: 4px;
    cursor: pointer;
    font-size: 16px;
    text-decoration: none;
    text-align: center;
  }}
  .btn:hover {{
    background-color: #45a049;
  }}
  .btn-reset {{
    background-color: #f44336;
  }}
  .btn-reset:hover {{
    background-color: #da190b;
  }}
  .btn-start {{
    background-color: #2196F3;
  }}
  .btn-start:hover {{
    background-color: #0b7dda;
  }}
  .btn-stop {{
    background-color: #ff9800;
  }}
  .btn-stop:hover {{
    background-color: #e68a00;
  }}
  .btn-record {{
    background-color: #9C27B0;
  }}
  .btn-record:hover {{
    background-color: #7B1FA2;
  }}
  .label {{
    text-align: center;
    margin-top: 15px;
    font-weight: bold;
    color: #666;
    font-size: 18px;
  }}
  .status {{
    text-align: center;
    margin: 10px 0;
    font-size: 18px;
    font-weight: bold;
    color: #e91e63;
  }}
  .record-table {{
    width: 100%;
    margin-top: 20px;
    border: 1px solid #ddd;
    border-collapse: collapse;
  }}
  .record-table th, .record-table td {{
    padding: 10px;
    text-align: left;
    border: 1px solid #ddd;
  }}
  .record-table th {{
    background-color: #f2f2f2;
  }}
  </style>
</head>
<body>
  <div class="container">
  <h2>签到情况</h2>
  {table_html}
  <div class="label">讲台</div>
  <div class="status">{status_text}</div>
  <form method="POST" action="/checkin/{classroom_id}/save">
    <div class="form-group">
      <label for="course">课程名称:</label>
      <input type="text" id="course" name="course" required>
    </div>
    <button type="submit" class="btn">保存</button>
    <button type="submit" class="btn btn-record" formaction="/checkin/{classroom_id}/view-records">签到记录</button>
  </form>

  {control_buttons}
  <!-- 添加新按钮 -->
  <form method="GET" action="/checkin/{classroom_id}/view-by-student">
    <button type="submit" class="btn btn-view">按学号查看签到情况</button>
  </form>
  <form method="POST" action="/checkin/{classroom_id}/reset" onsubmit="return confirm('确定要重置所有数据吗？此操作不可恢复！');">
    <input type="hidden" name="action" value="reset">
    <button type="submit" class="btn btn-reset">重置</button>
  </form>
  </div>
</body>
</html>'''

    def _render_admin(self, table_html='', classroom_id=''):
        """动态生成 admin 页面"""
        # 判断签到状态
        is_checkin_active = CheckinHandler.checkin_enabled.get(classroom_id, False)
        status_text = "正在签到..." if is_checkin_active else "未开始签到"
        
        # 生成控制按钮
        if is_checkin_active:
            control_buttons = '''
  <form method="POST" action="/checkin/{classroom_id}/stop-checkin">
    <button type="submit" class="btn btn-stop">结束签到</button>
  </form>'''.format(classroom_id=classroom_id)
        else:
            control_buttons = '''
  <form method="POST" action="/checkin/{classroom_id}/start-checkin">
    <button type="submit" class="btn btn-start">开始签到</button>
  </form>'''.format(classroom_id=classroom_id)


        html = self._admin_template.format(
            table_html=table_html,
            classroom_id=classroom_id,
            status_text=status_text,
            control_buttons=control_buttons
        )
        return html.encode('utf-8')

    def _render_form(self, message=''):
        try:
            html_content = importlib.resources.read_text('checkin', 'checkin.html', encoding='utf-8')
        except FileNotFoundError:
            html_content = "<html><body><h2>页面丢失</h2></body></html>"
        except Exception:
            html_content = "<html><body><h2>模板加载失败</h2></body></html>"

        msg_html = f'<p style="color:green">{message}</p>' if message else ''
        html_content = html_content.replace('{{message}}', msg_html)
        return html_content.encode('utf-8')

    def _render_manage(self):
        try:
            return importlib.resources.read_text('checkin', 'manage.html', encoding='utf-8').encode('utf-8')
        except Exception:
            return b"<h2>Manage template missing</h2>"

    def _get_room_config(self, classroom_id):
        """从数据库获取教室信息"""
        print(f"[DEBUG] Looking for classroom_id: {classroom_id}")
        result = get_classroom_by_id(classroom_id)
        if result:
            print(f"[DEBUG] Found room config: {result}")
            return result
        print(f"[DEBUG] Classroom {classroom_id} not found")
        return (None, None, None)

    def _generate_qr_codes(self, classroom_id):
        """生成指定教室的二维码"""
        classroom_id, row, col = self._get_room_config(classroom_id)
        if not classroom_id:
            return False
            
        public_ip = CheckinHandler.public_ip
        total_seats = min(row * col, 48)
        
        # 创建输出目录
        output_dir = os.path.join("data", classroom_id, "qrcode")
        os.makedirs(output_dir, exist_ok=True)
        
        base_url = f"http://{public_ip}/checkin/{classroom_id}/checkin-{{:02d}}.html"
        
        for num in range(1, total_seats + 1):
            url = base_url.format(num)
            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=10,
                border=4,
            )
            qr.add_data(url)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white").convert("RGB")
            
            draw = ImageDraw.Draw(img)
            try:
                font = ImageFont.truetype("arial.ttf", 20)
            except IOError:
                font = ImageFont.load_default()
            
            text = f"{num:02d}"
            text_bbox = draw.textbbox((0, 0), text, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            img_width, _ = img.size
            position = ((img_width - text_width) // 2, 5)
            draw.text(position, text, font=font, fill="black")
            
            filename = os.path.join(output_dir, f"qr-{num:02d}.png")
            img.save(filename)
            
        return True

    def _generate_latex_file(self, classroom_id):
        """生成 LaTeX 文件用于打印二维码"""
        classroom_id, row, col = self._get_room_config(classroom_id)
        if not classroom_id:
            return None
            
        total_seats = min(row * col, 48)
        output_dir = os.path.join("data", classroom_id, "qrcode")
        
        # 检查二维码文件是否存在
        for i in range(1, total_seats + 1):
            qr_file = os.path.join(output_dir, f"qr-{i:02d}.png")
            if not os.path.exists(qr_file):
                return None
        
        # 生成 LaTeX 内容
        latex_content = r"""\documentclass[a4paper,10pt]{article}
\usepackage[margin=1cm]{geometry}
\usepackage{graphicx}
\usepackage{caption}
\usepackage{subcaption}
\usepackage{tikz}

\setlength{\parindent}{0pt}
\pagestyle{empty} 

\begin{document}

"""
        
        # 添加二维码包含命令
        for i in range(1, total_seats + 1):
            latex_content += f"  \\includegraphics[width=0.23\\textwidth]{{qr-{i:02d}.png}}%\n"
            if i < total_seats:
                remainder = (i - 1) % 4
                if remainder == 3:
                    latex_content += "  \\par\n"
                else:
                    latex_content += "  \\hfill\n"
        
        latex_content += r"\end{document}"
        
        # 保存 LaTeX 文件
        tex_file = os.path.join(output_dir, f"qrcode-{classroom_id}.tex")
        with open(tex_file, 'w', encoding='utf-8') as f:
            f.write(latex_content)
            
        return tex_file

    def _compile_latex_to_pdf(self, tex_file_path):
        """调用 pdflatex 编译 LaTeX 文件为 PDF"""
        try:
            # 获取 LaTeX 文件所在目录
            tex_dir = os.path.dirname(tex_file_path)
            tex_filename = os.path.basename(tex_file_path)
            
            # 在 LaTeX 文件目录中执行 pdflatex
            result = subprocess.run(
                ['pdflatex', '-interaction=nonstopmode', tex_filename],
                cwd=tex_dir,
                capture_output=True,
                text=True,
                timeout=30  # 30秒超时
            )
            
            if result.returncode == 0:
                pdf_file = tex_file_path.replace('.tex', '.pdf')
                if os.path.exists(pdf_file):
                    return pdf_file
            else:
                print(f"pdflatex error: {result.stderr}")
                return None
                
        except subprocess.TimeoutExpired:
            print("pdflatex timeout")
            return None
        except FileNotFoundError:
            print("pdflatex not found. Please install LaTeX distribution.")
            return None
        except Exception as e:
            print(f"Error compiling LaTeX: {e}")
            return None

    def _build_table_html(self, classroom_id):
        """基于内存配置构建表格"""
        classroom_id, row, col = self._get_room_config(classroom_id)  # ✅ 接收 id
        if not classroom_id:
            return "<h2>配置错误</h2>"
        
        row = row or 4
        col = col or 12
        
        # 从 checkin-temp 表读取签到数据
        from .database import get_temp_checkins_by_classroom
        temp_checkins = get_temp_checkins_by_classroom(classroom_id)
        
        # 构建表格
        table = [["" for _ in range(col)] for _ in range(row)]
        for name, seat_number in temp_checkins:
            try:
                idx = seat_number - 1
                r = idx // col
                c = idx % col
                if 0 <= r < row and 0 <= c < col:
                    table[r][c] = name
            except (ValueError, IndexError):
                continue

        # 生成HTML (倒序显示行)
        table_html = "<table border='1' style='width:100%; border-collapse: collapse;'>\n"
        for tr in reversed(table):  # 倒序显示
            table_html += "  <tr>\n"
            for cell in tr:
                table_html += f"    <td>{cell}</td>\n"
            table_html += "  </tr>\n"
        table_html += "</table>"
        return table_html
    
    def do_GET(self):
        path = urllib.parse.urlparse(self.path).path

        # ✅ 修改路由: /checkin/manage.html
        if path == "/checkin/manage.html":
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(self._render_manage())
            return

        # 新增：提供 qrcode 目录下的静态文件（PDF/PNG）
        qr_match = re.match(r'^/checkin/(\d{3,4})/qrcode/(.+)$', path)
        if qr_match:
            classroom_id = qr_match.group(1)
            filename = qr_match.group(2)
            
            # 安全校验：只允许 .pdf 和 .png
            if not (filename.endswith('.pdf') or filename.endswith('.png')):
                self.send_response(403)
                self.end_headers()
                self.wfile.write(b"Forbidden: Invalid file type")
                return

            file_path = os.path.join("data", classroom_id, "qrcode", filename)
            if os.path.exists(file_path) and os.path.isfile(file_path):
                self.send_response(200)
                if filename.endswith('.pdf'):
                    self.send_header('Content-Type', 'application/pdf')
                else:
                    self.send_header('Content-Type', 'image/png')
                self.send_header('Content-Disposition', f'inline; filename="{filename}"')
                self.end_headers()
                with open(file_path, 'rb') as f:
                    self.wfile.write(f.read())
            else:
                self.send_response(404)
                self.end_headers()
                self.wfile.write(b"<h2>File not found</h2>")
            return

        # 新增：列出所有教室
        if path == "/checkin/manage/list":
            classrooms = get_all_classrooms()
            public_ip = getattr(CheckinHandler, 'public_ip', 'localhost')
            
            html = "<!DOCTYPE html><html><head><meta charset='utf-8'><title>教室列表</title></head><body>"
            html += "<h2>当前配置的教室</h2>"
            html += f"<p><strong>公共IP:</strong> {public_ip}</p>"
            html += "<ul>"
            for room in classrooms:
                html += f"<li>教室ID: {room['id']}, 行: {room['row']}, 列: {room['column']} "
                html += f'<a href="/checkin/{room["id"]}/admin.html" style="margin-left:10px;">查看教室签到情况</a></li>'
            html += "</ul>"
            html += '<p><a href="/checkin/manage.html">返回管理页面</a></p>'
            html += "</body></html>"
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        # ✅ 匹配 /checkin/{id}/admin.html 或 /checkin/{id}/checkin-XX.html
        match = re.match(r'^/checkin/(\d{3,4})/(admin\.html|checkin-\d{2}\.html)$', path)
        if match:
            classroom_id = match.group(1)
            page_type = match.group(2)

            # 使用内存配置替代文件加载
            classroom_id, _, _ = self._get_room_config(classroom_id)  # ✅ 接收 id
            if classroom_id is None:
                self.send_response(404)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write("<h2>教室配置未找到</h2>".encode('utf-8'))
                return

            if page_type == "admin.html":
                # 构建表格时也使用内存配置
                table_html = self._build_table_html(classroom_id)
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(self._render_admin(table_html=table_html, classroom_id=classroom_id))  # ✅ 传递 classroom_id
                return

            elif page_type.startswith("checkin-"):
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(self._render_form())
                return

        # 新增：列出已导入的班级及学生数量
        if path == "/checkin/manage/list-students":
            classes = get_class_student_counts()
            
            html = '<div style="font-family: Arial, sans-serif;">'
            if not classes:
                html += "<p>暂无导入的班级数据</p>"
            else:
                for cls in classes:
                    html += f'''
                    <div class="class-item">
                        <span><strong>{cls["class"]}</strong> ({cls["count"]} 名学生)</span>
                        <button class="delete-btn" onclick="parent.deleteClass('{cls["class"]}')">删除班级</button>
                    </div>
                    '''
            html += '</div>'
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        # 新增：返回导入学生页面（动态生成）
        if path == "/checkin/import-student.html":
            html = '''<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>导入学生名单</title>
  <style>
    body { font-family: sans-serif; padding: 20px; }
    .form-group { margin: 15px 0; }
    input, button { padding: 8px; }
    button { background: #4CAF50; color: white; border: none; cursor: pointer; }
  </style>
</head>
<body>
  <h2>导入班级学生名单</h2>
  <form method="POST" action="/checkin/manage/import-students" enctype="multipart/form-data">
    <div class="form-group">
      <label for="csv_file">选择 CSV 文件：</label>
      <input type="file" id="csv_file" name="csv_file" accept=".csv" required>
    </div>
    <button type="submit">上传并导入</button>
  </form>
  <p><a href="/checkin/manage.html">返回管理页面</a></p>
</body>
</html>'''
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        # 新增：按学号查看签到情况
        student_view_match = re.match(r'^/checkin/(\d{3,4})/view-by-student$', path)
        if student_view_match:
            classroom_id = student_view_match.group(1)
            
            # 获取该教室对应的班级名称
            from .database import get_class_name_by_classroom, get_students_by_class_name, get_temp_checkins_with_ids_by_classroom, get_classroom_by_id
            class_name = get_class_name_by_classroom(classroom_id)
            all_students = get_students_by_class_name(class_name)

            # 获取教室配置以确定最大座位数
            classroom_config = get_classroom_by_id(classroom_id)
            if classroom_config:
                _, row, col = classroom_config
                max_seats = min(row * col, 48)
            else:
                max_seats = 48
            
            # 获取该教室的临时签到数据（包含学号和状态）
            temp_checkins = get_temp_checkins_with_ids_by_classroom(classroom_id)
            
            # 创建签到状态字典和座位号字典（使用学号作为键），默认状态改为"缺勤"
            checkin_status = {student_id: "缺勤" for student_id, _ in all_students}
            seat_numbers = {student_id: "-" for student_id, _ in all_students}  # 默认座位号为"-"
            for student_id, _, seat_num, status in temp_checkins:  # 现在包含座位号
                if student_id in checkin_status:
                    checkin_status[student_id] = status
                    # 只有"已签"状态才显示实际座位号，其他状态显示"-"
                    if status == "已签" and seat_num:
                        seat_numbers[student_id] = str(seat_num)
                    else:
                        seat_numbers[student_id] = "-"
            
            # 生成HTML表格（无JavaScript）
            table_html = "<table border='1' style='width:100%; border-collapse: collapse;'>"
            table_html += "<tr><th>学号</th><th>姓名</th><th>签到状态</th><th>座位号</th></tr>"  # 删除了 <th>操作</th>
        
            for student_id, name in all_students:
                status = checkin_status.get(student_id, "缺勤")
                seat_num = seat_numbers.get(student_id, "-")
                table_html += f"""
            <tr>
                <td>{student_id}</td>
                <td>{name}</td>
                <td>
                    <select name="status_{student_id}">
                        <option value="已签"{" selected" if status == "已签" else ""}>已签</option>
                        <option value="缺勤"{" selected" if status == "缺勤" else ""}>缺勤</option>
                        <option value="病假"{" selected" if status == "病假" else ""}>病假</option>
                        <option value="事假"{" selected" if status == "事假" else ""}>事假</option>
                        <option value="公假"{" selected" if status == "公假" else ""}>公假</option>
                        <option value="迟到"{" selected" if status == "迟到" else ""}>迟到</option>
                        <option value="早退"{" selected" if status == "早退" else ""}>早退</option>
                    </select>
                </td>
                <td>
                    <input type="text" name="seat_{student_id}" value="{seat_num}" style="width:60px;">
                </td>
                <!-- 删除了操作列 <td>...</td> -->
            </tr>"""
        
            table_html += "</table>"
        
            # 生成完整页面（无JavaScript）
            html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>按学号查看签到情况</title>
  <style>
    body {{ font-family: sans-serif; padding: 20px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
    th, td {{ padding: 10px; text-align: left; border: 1px solid #ddd; }}
    th {{ background-color: #f2f2f2; }}
    .btn {{ display: inline-block; margin-top: 20px; padding: 10px 20px; 
           background-color: #4CAF50; color: white; text-decoration: none; border-radius: 4px; }}
    .btn:hover {{ background-color: #45a049; }}
    .save-btn {{ background-color: #2196F3; }}
    .save-btn:hover {{ background-color: #0b7dda; }}
    input[type="text"] {{
        width: 60px;
        padding: 5px;
        border: 1px solid #ccc;
        border-radius: 4px;
    }}
  </style>
</head>
<body>
  <h2>教室 {classroom_id} 学生签到情况</h2>
  <form method="POST" action="/checkin/{classroom_id}/update-student-status">
    {table_html}
    <button type="submit" class="btn save-btn">保存更改</button>
  </form>
  <a href="/checkin/{classroom_id}/admin.html" class="btn">返回管理页面</a>
</body>
</html>"""
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        self.send_response(404)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.end_headers()
        self.wfile.write("<h2>无效路径，请通过 /checkin/{教室ID}/admin.html 访问</h2>".encode('utf-8'))

    def do_POST(self):
        path = urllib.parse.urlparse(self.path).path

        # 新增：删除签到记录
        if path == "/checkin/delete-record":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            course = params.get("course", [""])[0]
            save_time = params.get("save_time", [""])[0]
            classroom_id = params.get("classroom_id", [""])[0]
            
            # 调用数据库函数删除记录
            from .database import delete_checkin_record
            if delete_checkin_record(course, save_time, classroom_id):
                # 重新查询记录以刷新页面
                from .database import get_checkin_summary_by_course
                records = get_checkin_summary_by_course(course)
                
                # 重新渲染页面
                if records:
                    table_rows = ""
                    for record in records:
                        table_rows += f"""
                    <tr>
                        <td>{record['course']}</td>
                        <td>{record['classroom_id']}</td>
                        <td>{record['count']}</td>
                        <td>{record['save_time']}</td>
                        <td>
                            <form method="POST" action="/checkin/delete-record" style="display:inline;">
                                <input type="hidden" name="course" value="{record['course']}">
                                <input type="hidden" name="save_time" value="{record['save_time']}">
                                <input type="hidden" name="classroom_id" value="{record['classroom_id']}">
                                <button type="submit" class="btn-delete">删除记录</button>
                            </form>
                        </td>
                    </tr>"""
                    table_html = f"""
                <table class="record-table">
                    <tr>
                        <th>课程名称</th>
                        <th>教室ID</th>
                        <th>签到人数</th>
                        <th>保存时间</th>
                        <th>操作</th>
                    </tr>
                    {table_rows}
                </table>"""
                else:
                    table_html = "<p>未找到相关签到记录</p>"
                
                html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>签到记录</title>
<style>
body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    margin: 0;
    padding: 20px;
    background-color: #f5f5f5;
}}
.container {{
    max-width: 800px;
    margin: 0 auto;
    background-color: white;
    padding: 20px;
    border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
}}
.record-table {{
    width: 100%;
    margin-top: 20px;
    border: 1px solid #ddd;
    border-collapse: collapse;
}}
.record-table th, .record-table td {{
    padding: 10px;
    text-align: left;
    border: 1极线 solid #ddd;
}}
.record-table th {{
    background-color: #f2f2f2;
}}
.btn {{
    display: inline-block;
    margin-top: 20px;
    padding: 10px 20px;
    background-color: #4CAF50;
    color: white;
    text-decoration: none;
    border-radius: 4px;
}}
.btn:hover {{
    background-color: #45a049;
}}
.btn-delete {{
    padding: 5px 10px; 
    background: #f44336; 
    color: white; 
    border: none; 
    cursor: pointer; 
    border-radius: 4px;
}}
</style>
</head>
<body>
  <div class="container">
    <h2>签到记录</h2>
    {table_html}
    <a href="/checkin/{classroom_id}/admin.html" class="btn">返回管理页面</a>
  </div>
</body>
</html>"""
            else:
                html_resp = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>删除失败</title></head>
<body>
<h2>删除失败</h2>
<p>记录删除失败，请重试。</p>
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # ✅ 查看签到记录
        record_match = re.match(r'^/checkin/(\d{3,4})/view-records$', path)
        if record_match:
            classroom_id = record_match.group(1)
            # 获取课程名称
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            course_name = params.get("course", [""])[0]
            
            if not course_name:
                html_resp = """<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>查询失败</title></head>
<body><p>请输入课程名称</p><p><a href="/checkin/{classroom_id}/admin.html">返回</a></p></body></html>""".format(classroom_id=classroom_id)
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(html_resp.encode('utf-8'))
                return
            
            # 查询签到记录
            from .database import get_checkin_summary_by_course
            records = get_checkin_summary_by_course(course_name)
            
            # 生成记录表格
            if records:
                # 使用复选框选择多条记录并由同一按钮导出
                table_rows = ""
                for record in records:
                    # 复选框的值编码为 course||save_time||classroom_id
                    cbval = f"{record['course']}||{record['save_time']}||{record['classroom_id']}"
                    table_rows += f"""
                <tr>
                    <td><input type="checkbox" name="export_record" value="{cbval}"></td>
                    <td>{record['course']}</td>
                    <td>{record['classroom_id']}</td>
                    <td>{record['class_total']}</td>
                    <td>{record['signed']}</td>
                    <td>{record['personal_leave']}</td>
                    <td>{record['sick_leave']}</td>
                    <td>{record['official_leave']}</td>
                    <td>{record['absent']}</td>
                    <td>{record['late']}</td>
                    <td>{record['early_leave']}</td>
                    <td>{record['save_time']}</td>
                    <td>
                        <form method="POST" action="/checkin/delete-record" style="display:inline;">
                            <input type="hidden" name="course" value="{record['course']}">
                            <input type="hidden" name="save_time" value="{record['save_time']}">
                            <input type="hidden" name="classroom_id" value="{record['classroom_id']}">
                            <button type="submit" class="btn-delete">删除记录</button>
                        </form>
                    </td>
                </tr>"""
                # 表格被包裹在一个表单内，表单提交时会发送所有被选中的 export_record 值
                table_html = f"""
            <form method="POST" action="/checkin/export-record">
            <table class="record-table">
                <tr>
                    <th>选择</th>
                    <th>课程名称</th>
                    <th>教室ID</th>
                    <th>班级人数</th>
                    <th>已签</th>
                    <th>事假</th>
                    <th>病假</th>
                    <th>公假</th>
                    <th>缺勤</th>
                    <th>迟到</th>
                    <th>早退</th>
                    <th>保存时间</th>
                    <th>操作</th>
                </tr>
                {table_rows}
            </table>
            <div style="margin-top:12px;">
                <button type="submit" class="btn-export">导出到xlsx文件</button>
            </div>
            </form>"""
            else:
                table_html = "<p>未找到相关签到记录</p>"
 
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>签到记录</title>
<style>
body {{
    font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    margin: 0;
    padding: 20px;
    background-color: #f5f5f5;
}}
.container {{
    max-width: 800px;
    margin: 0 auto;
    background-color: white;
    padding: 20px;
    border-radius: 8px;
    box-shadow: 0 2px 10px rgba(0,0,0,0.1);
}}
.record-table {{
    width: 100%;
    margin-top: 20px;
    border: 1px solid #ddd;
    border-collapse: collapse;
}}
.record-table th, .record-table td {{
    padding: 10px;
    text-align: left;
    border: 1px solid #ddd;
}}
.record-table th {{
    background-color: #f2f2f2;
}}
.btn {{
    display: inline-block;
    margin-top: 20px;
    padding: 10px 20px;
    background-color: #4CAF50;
    color: white;
    text-decoration: none;
    border-radius: 4px;
}}
.btn:hover {{
    background-color: #45a049;
}}
.btn-delete {{
    padding: 5px 10px; 
    background: #f44336; 
    color: white; 
    border: none; 
    cursor: pointer; 
    border-radius: 4px;
}}
</style>
</head>
<body>
  <div class="container">
    <h2>签到记录</h2>
    {table_html}
    <a href="/checkin/{classroom_id}/admin.html" class="btn">返回管理页面</a>
  </div>
</body>
</html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # ✅ 添加教室: /checkin/manage/add
        if path == "/checkin/manage/add":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            classroom_id = params.get("classroom_id", [""])[0]
            row = int(params.get("row", ["4"])[0])
            column = int(params.get("column", ["12"])[0])
            
            add_classroom(classroom_id, row, column)
            
            self.send_response(302)
            self.send_header('Location', "/checkin/manage/list")
            self.end_headers()
            return

        # ✅ 删除教室: /checkin/manage/delete
        if path == "/checkin/manage/delete":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            classroom_id_to_delete = params.get("classroom_id", [""])[0]
            delete_classroom(classroom_id_to_delete)
            
            self.send_response(302)
            self.send_header('Location', "/checkin/manage.html")
            self.end_headers()
            return
    
            # ✅ 生成二维码: /checkin/manage/generate-qrcode
        if path == "/checkin/manage/generate-qrcode":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            classroom_id = params.get("classroom_id", [""])[0]
            
            if self._generate_qr_codes(classroom_id):
                message = f"二维码已生成到 ./data/{classroom_id}/qrcode/ 目录"
                # 添加下载按钮
                download_button = f'<form method="POST" action="/checkin/manage/generate-print-file" style="margin-top: 15px;">' \
                                f'<input type="hidden" name="classroom_id" value="{classroom_id}">' \
                                f'<button type="submit" class="btn-qrcode">下载打印文件</button>' \
                                f'</form>'
            else:
                message = "教室ID不存在，无法生成二维码"
                download_button = ""
            
            # 返回结果页面
            html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>二维码生成结果</title>
<style>
.btn-qrcode {{ 
    padding: 10px 20px; 
    background: #9C27B0; 
    color: white; 
    border: none; 
    cursor: pointer; 
    margin-top: 10px;
}}
</style>
</head>
<body>
<h2>二维码生成结果</h2>
<p>{message}</p>
{download_button}
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        # 新增：生成打印文件（LaTeX + PDF）
        if path == "/checkin/manage/generate-print-file":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            classroom_id = params.get("classroom_id", [""])[0]
            
            # 生成 LaTeX 文件
            tex_file = self._generate_latex_file(classroom_id)
            if tex_file:
                # 编译为 PDF
                pdf_file = self._compile_latex_to_pdf(tex_file)
                if pdf_file:
                    # 重定向到下载页面
                    download_url = f"/checkin/{classroom_id}/qrcode/qrcode-{classroom_id}.pdf"
                    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>打印文件生成成功</title></head>
<body>
<h2>打印文件生成成功</h2>
<p>PDF文件已生成，点击下面链接下载:</p>
<p><a href="{download_url}" style="font-size: 18px; color: #2196F3;">下载 qrcode-{classroom_id}.pdf</a></p>
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
                else:
                    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>PDF生成失败</title></head>
<body>
<h2>PDF生成失败</h2>
<p>LaTeX文件已生成，但编译PDF失败。请确保已安装LaTeX发行版（如MiKTeX或TeX Live）。</p>
<p>LaTeX文件位置: {tex_file}</p>
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
            else:
                html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>LaTeX生成失败</title></head>
<body>
<h2>LaTeX生成失败</h2>
<p>无法生成LaTeX文件。请确保已先生成二维码。</p>
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
            
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html.encode('utf-8'))
            return

        # 新增：导入学生名单
        if path == "/checkin/manage/import-students":
            content_type = self.headers.get('Content-Type', '')
            if 'multipart/form-data' not in content_type:
                self._send_import_result("无效的请求类型", success=False)
                return

            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)

            # 提取 boundary
            boundary_match = re.search(r'boundary=([^;]+)', content_type)
            if not boundary_match:
                self._send_import_result("无效的 multipart 格式", success=False)
                return
            boundary = boundary_match.group(1).strip('"').encode()

            parts = body.split(b'--' + boundary)
            csv_content = None
            filename = None
            for part in parts:
                if b'name="csv_file"' in part:
                    # 提取文件名
                    filename_match = re.search(rb'filename="([^"]+)"', part)
                    if filename_match:
                        filename = filename_match.group(1).decode('utf-8')
                    header_end = part.find(b'\r\n\r\n')
                    if header_end != -1:
                        csv_content = part[header_end+4:]
                        if csv_content.endswith(b'\r\n'):
                            csv_content = csv_content[:-2]
                        break

            if not filename or not csv_content:
                self._send_import_result("未选择文件", success=False)
                return

            try:
                import csv
                from io import StringIO
                decoded = csv_content.decode('utf-8-sig')
                reader = csv.reader(StringIO(decoded))
                students = []
                seen_ids = set()
                duplicates = []
                for row in reader:
                    if len(row) >= 3:
                        student_id, name, class_name = row[0].strip(), row[1].strip(), row[2].strip()
                        if not student_id or not name or not class_name:
                            continue
                        if student_id in seen_ids:
                            duplicates.append(student_id)
                        else:
                            seen_ids.add(student_id)
                            students.append((student_id, name, class_name))
                
                if duplicates:
                    self._send_import_result(f"导入失败：发现重复学号 {', '.join(duplicates)}", success=False)
                    return

                if not students:
                    self._send_import_result(f"文件 '{filename}' 为空或格式不正确", success=False)
                    return

                conn = sqlite3.connect(DATABASE_PATH)
                cursor = conn.cursor()
                try:
                    cursor.executemany(
                        "INSERT INTO students (student_id, name, class_name) VALUES (?, ?, ?)",
                        students
                    )
                    conn.commit()
                    self._send_import_result(f"成功导入 '{filename}' 中的 {len(students)} 名学生")
                except sqlite3.IntegrityError as e:
                    if "UNIQUE constraint failed" in str(e):
                        conn.rollback()
                        # 查询哪些学号已存在
                        placeholders = ','.join('?' for _ in seen_ids)
                        cursor.execute(f"SELECT student_id FROM students WHERE student_id IN ({placeholders})", tuple(seen_ids))
                        existing_ids = [row[0] for row in cursor.fetchall()]
                        conn.close()
                        self._send_import_result(f"导入失败：以下学号已存在 {', '.join(existing_ids)}", success=False)
                    else:
                        raise
                finally:
                    conn.close()

            except Exception as e:
                self._send_import_result(f"导入失败: {str(e)}", success=False)
            return

        # 新增：删除指定班级的学生名单
        if path == "/checkin/manage/delete-class-students":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            class_name = params.get("class_name", [""])[0]
            if not class_name:
                self._send_import_result("班级名称不能为空", success=False)
                return

            deleted_count = delete_students_by_class_name(class_name)
            
            if deleted_count > 0:
                message = f"成功删除班级 '{class_name}' 中的 {deleted_count} 名学生"
            else:
                message = f"班级 '{class_name}' 不存在或无学生可删除"
                
            self._send_import_result(message)
            return

        # 新增：更新学生签到状态
        update_status_match = re.match(r'^/checkin/(\d{3,4})/update-student-status$', path)
        if update_status_match:
            classroom_id = update_status_match.group(1)
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            
            # 获取该教室对应的班级名称
            from .database import get_class_name_by_classroom, get_students_by_class_name, clear_temp_checkins, add_temp_checkin, get_classroom_by_id
            class_name = get_class_name_by_classroom(classroom_id)
            all_students = get_students_by_class_name(class_name)

            # 获取教室配置以确定最大座位数
            classroom_config = get_classroom_by_id(classroom_id)
            if classroom_config:
                _, row, col = classroom_config
                max_seats = min(row * col, 48)
            else:
                max_seats = 48
            
            # 验证所有输入
            validation_errors = []
            for student_id, name in all_students:
                status_key = f"status_{student_id}"
                seat_key = f"seat_{student_id}"
                
                status_value = params.get(status_key, ["缺勤"])[0]
                seat_input = params.get(seat_key, ["-"])[0].strip()
                
                if status_value == "已签":
                    if seat_input == "" or seat_input == "-":
                        validation_errors.append(f"学号 {student_id}（{name}）：座位号不能为空")
                    else:
                        try:
                            seat_num_val = int(seat_input)
                            if seat_num_val < 1 or seat_num_val > max_seats:
                                validation_errors.append(f"学号 {student_id}（{name}）：座位号 {seat_num_val} 超出范围（1-{max_seats}）")
                        except ValueError:
                            validation_errors.append(f"学号 {student_id}（{name}）：座位号 '{seat_input}' 不是有效数字")
            
            if validation_errors:
                # 返回错误页面
                error_html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>输入错误</title>
<style>
body {{ font-family: sans-serif; padding: 20px; color: #d32f2f; }}
ul {{ margin-top: 10px; }}
.btn {{ display: inline-block; margin-top: 20px; padding: 10px 20px; 
       background-color: #4CAF50; color: white; text-decoration: none; border-radius: 4px; }}
</style>
</head>
<body>
<h2>保存失败：发现以下错误</h2>
<ul>{''.join(f'<li>{err}</li>' for err in validation_errors)}</ul>
<a href="/checkin/{classroom_id}/view-by-student" class="btn">返回修改</a>
</body></html>"""
                self.send_response(400)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(error_html.encode('utf-8'))
                return
            
            # 清空当前教室的所有临时签到记录
            clear_temp_checkins(classroom_id)
            
            # 重新添加所有学生记录
            updated_count = 0
            for student_id, name in all_students:
                status_value = params.get(f"status_{student_id}", ["缺勤"])[0]
                seat_input = params.get(f"seat_{student_id}", ["-"])[0].strip()
                
                seat_num = None
                if status_value == "已签":
                    seat_num = int(seat_input)  # 已通过验证
                
                if add_temp_checkin(student_id, classroom_id, seat_num, status_value):
                    updated_count += 1
            
            redirect_url = f"/checkin/{classroom_id}/view-by-student"
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>更新成功</title>
<meta http-equiv="refresh" content="2;url={redirect_url}"></head>
<body><p>已更新 {updated_count} 名学生的签到状态，2秒后返回...</p></body></html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # ✅ 匹配 /checkin/{id}/save 和 /checkin/{id}/reset
        save_match = re.match(r'^/checkin/(\d{3,4})/save$', path)
        reset_match = re.match(r'^/checkin/(\d{3,4})/reset$', path)

        if save_match:
            classroom_id = save_match.group(1)
            # 获取课程名称
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)
            course_name = params.get("course", [""])[0]
            
            # 保存到数据库
            from .database import save_checkin_records
            count = save_checkin_records(classroom_id, course_name)
            
            redirect_url = f"/checkin/{classroom_id}/admin.html"
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>保存成功</title>
<meta http-equiv="refresh" content="2;url={redirect_url}"></head>
<body><p>已保存 {count} 条签到记录，2秒后返回...</p></body></html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        if reset_match:
            classroom_id = reset_match.group(1)
            admin.reset_namefile(classroom_id=classroom_id)
            redirect_url = f"/checkin/{classroom_id}/admin.html"
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>重置成功</title>
<meta http-equiv="refresh" content="1;url={redirect_url}"></head>
<body><p>数据已重置，1秒后返回...</p></body></html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # ✅ 开始签到
        start_match = re.match(r'^/checkin/(\d{3,4})/start-checkin$', path)
        if start_match:
            classroom_id = start_match.group(1)
            CheckinHandler.checkin_enabled[classroom_id] = True
            redirect_url = f"/checkin/{classroom_id}/admin.html"
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>签到已开始</title>
<meta http-equiv="refresh" content="1;url={redirect_url}"></head>
<body><p>签到已开始，学生可以扫码签到。</p></body></html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # ✅ 结束签到
        stop_match = re.match(r'^/checkin/(\d{3,4})/stop-checkin$', path)
        if stop_match:
            classroom_id = stop_match.group(1)
            CheckinHandler.checkin_enabled[classroom_id] = False
            redirect_url = f"/checkin/{classroom_id}/admin.html"
            html_resp = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>签到已结束</title>
<meta http-equiv="refresh" content="1;url={redirect_url}"></head>
<body><p>签到已结束，学生无法继续签到。</p></body></html>"""
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(html_resp.encode('utf-8'))
            return

        # 仅在路径为 /checkin/{id}/checkin-XX.html 时处理学生扫码签到请求，
        # 否则保留给其它 POST 分支（比如导出记录）处理。
        checkin_post_match = re.match(r'^/checkin/(\d{3,4})/checkin-(\d{2})\.html$', path)
        if checkin_post_match:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length) if content_length > 0 else b''
            content_type = self.headers.get('Content-Type', '')

            student_id = None
            if 'application/json' in content_type:
                try:
                    data = json.loads(body.decode('utf-8'))
                    student_id = data.get('student_id') or data.get('user_id')
                except json.JSONDecodeError:
                    student_id = None
            else:
                try:
                    parsed = urllib.parse.parse_qs(body.decode('utf-8'))
                    student_id = parsed.get('student_id', [None])[0]
                except Exception:
                    student_id = None

            if student_id:
                classroom_id = checkin_post_match.group(1)
                seq = int(checkin_post_match.group(2))

                # 检查是否允许签到
                if not CheckinHandler.checkin_enabled.get(classroom_id, False):
                    message = "签到未开始或已结束"
                    status = 403
                else:
                    # 查询数据库获取姓名
                    conn = sqlite3.connect(DATABASE_PATH)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM students WHERE student_id = ?", (student_id,))
                    row = cursor.fetchone()
                    conn.close()

                    if not row:
                        message = "学号未找到，请确认是否已导入名单"
                        status = 400
                    else:
                        name = row[0]
                        from .database import add_temp_checkin
                        if add_temp_checkin(student_id, classroom_id, seq, "已签"):
                            message = f"签到成功：{name}"
                            status = 200
                        else:
                            message = "签到失败"
                            status = 500
            else:
                message = "缺少学号"
                status = 400

            self.send_response(status)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.end_headers()
            self.wfile.write(self._render_form(message=message))
            return
        

        if path == "/checkin/export-record":
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length).decode('utf-8')
            params = urllib.parse.parse_qs(body)

            # 支持批量导出：优先从表单的 export_record[] 获取多个选中项（格式为 course||save_time||classroom_id）
            export_items = params.get("export_record", [])

            rows = []
            # ensure these exist for filename/header fallback
            course = save_time = classroom_id = ""

            # 如果用户未选中任何复选框，直接返回提示页面，不做其它操作
            if not export_items:
                html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>未选择记录</title></head>
<body style="font-family: sans-serif; padding:20px;">
  <h2>没有选择任何签到记录</h2>
  <p>请返回并选择至少一条签到记录后再导出。</p>
</body></html>"""
                self.send_response(200)
                self.send_header('Content-Type', 'text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(html.encode('utf-8'))
                return
            
            if export_items:
                # 使用第一个选中项作为导出文件命名与表头日期的来源
                first_meta = None
                try:
                    from .database import get_checkin_records_by_save_time
                except Exception:
                    get_checkin_records_by_save_time = None

                for item in export_items:
                    try:
                        c, st, cid = item.split("||", 2)
                    except Exception:
                        continue
                    if first_meta is None:
                        first_meta = (c, st, cid)
                        course, save_time, classroom_id = c, st, cid

                    # 首选 helper 查询
                    recs = None
                    if get_checkin_records_by_save_time:
                        try:
                            recs = get_checkin_records_by_save_time(c, st, cid)
                        except Exception:
                            recs = None

                    # 回退到直连查询
                    if recs is None:
                        try:
                            conn = sqlite3.connect(DATABASE_PATH)
                            cur = conn.cursor()
                            cur.execute(
                                "SELECT student_id, name, status FROM checkin_records WHERE course=? AND save_time=? AND classroom_id=?",
                                (c, st, cid)
                            )
                            fetched = cur.fetchall()
                            conn.close()
                            recs = [{"student_id": r[0], "name": r[1], "status": r[2]} for r in fetched]
                        except Exception:
                            recs = []

                    for r in recs:
                        if isinstance(r, dict):
                            rows.append(r)
                        else:
                            rows.append({"student_id": r[0], "name": r[1], "status": r[2]})
            # end of export_items handling, continue with rows aggregation...

            if not rows:
                self._send_import_result("未找到符合条件的签到记录", success=False)
                return

            # 准备生成 xlsx
            try:
                from io import BytesIO
                import openpyxl
            except ImportError:
                self._send_import_result("服务器缺少 openpyxl 库，请 pip install openpyxl", success=False)
                return
            except Exception as e:
                self._send_import_result(f"无法初始化导出模块: {e}", success=False)
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            # 确保 ws 不为 None；如果为 None 或者设置 title 失败，则创建新 sheet
            if ws is None:
                ws = wb.create_sheet(title="签到明细")
            else:
                try:
                    ws.title = "签到明细"
                except Exception:
                    ws = wb.create_sheet(title="签到明细")
            # 移除多余的空默认 sheet（可选），保留名为 "签到明细" 的表
            try:
                if len(wb.sheetnames) > 1:
                    for name in list(wb.sheetnames):
                        if name != "签到明细":
                            sh = wb[name]
                            if sh.max_row == 1 and sh.max_column == 1 and sh.cell(1, 1).value is None:
                                wb.remove(sh)
            except Exception:
                # 如果移除失败也不影响后续导出
                pass

            # 小工具：从各种字符串中提取并标准化为 yyyy-mm-dd
            def _extract_date(val, fallback=""):
                if not val:
                    return fallback
                s = str(val)
                m = re.search(r'(\d{4}-\d{2}-\d{2})', s)
                if m:
                    return m.group(1)
                m = re.search(r'(\d{4}/\d{2}/\d{2})', s)
                if m:
                    return m.group(1).replace('/', '-')
                m = re.search(r'(\d{8})', s)
                if m:
                    g = m.group(1)
                    return f"{g[0:4]}-{g[4:6]}-{g[6:8]}"
                return fallback

            # 第三列标题为指定的日期（yyyy-mm-dd），优先使用请求中的 save_time，否则使用今天
            header_date = _extract_date(save_time, "")
            if not header_date:
                header_date = datetime.date.today().isoformat()
            ws.append(["学号", "姓名", header_date])

            # 列内容保留为签到状态（status）
            for r in rows:
                sid = name = status = ""
                if isinstance(r, dict):
                    sid = r.get("student_id") or r.get("id") or r.get("学号") or ""
                    name = r.get("name") or r.get("姓名") or ""
                    status = r.get("status") or r.get("state") or r.get("状态") or ""
                else:
                    vals = list(r)
                    if len(vals) >= 1:
                        sid = vals[0] or ""
                    if len(vals) >= 2:
                        name = vals[1] or ""
                    if len(vals) >= 3:
                        status = vals[2] or ""
                ws.append([sid, name, status])

            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            # 构造安全的文件名（移除潜在危险字符并限制长度）
            def _safe_name(s: str, maxlen=60):
                s = str(s or "")
                s = re.sub(r'[\\/:*?"<>|]+', "_", s)
                s = re.sub(r'\s+', "_", s)
                return s[:maxlen]

            # 当为批量导出时，若未提供明确 course/save_time/classroom_id，使用通用命名
            if export_items:
                now_tag = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                safe_course = "multiple"
                safe_time = now_tag
                safe_class = "multiple"
            else:
                safe_course = _safe_name(course)
                safe_time = _safe_name(save_time)
                safe_class = _safe_name(classroom_id)

            raw_fname = f"checkin_{safe_class}_{safe_course}_{safe_time}.xlsx"
            # 为 Content-Disposition 做 URL 引用，确保中文也可用
            quoted = urllib.parse.quote(raw_fname)

            # 构造仅包含 ASCII 的 header 值以避免 latin-1 编码错误：
            # 用不可打印/非 ASCII 字符替换为下划线作为 filename 回退，
            # 同时保留 RFC5987 的 filename*（使用 percent-encoding 的 UTF-8）。
            ascii_fname = re.sub(r'[^\x20-\x7E]', '_', raw_fname) or "download.xlsx"
            disposition = f'attachment; filename="{ascii_fname}"; filename*=UTF-8\'\'{quoted}'

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', disposition)
            self.end_headers()
            self.wfile.write(bio.read())
            return

    def _send_import_result(self, message, success=True):
        """返回导入结果页面"""
        status = 200 if success else 400
        html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>导入结果</title></head>
<body>
<h2>导入结果</h2>
<p>{message}</p>
<p><a href="/checkin/manage.html">返回管理页面</a></p>
</body></html>"""
        self.send_response(status)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.end_headers()
        self.wfile.write(html.encode('utf-8'))

